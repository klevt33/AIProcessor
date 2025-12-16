import asyncio
import logging
import sys

import openpyxl
from config_e2e import (
    DB_TABLE_NAME,
    EXCEL_FILE_PATH,
    EXCEL_TAB_NAME,
    PRIMARY_KEY_FIELD,
    EnhancedSDP,
    add_columns_to_right,
    clean_column_headers,
    clean_row_values,
)

from config import Config

# --- Configuration ---
BATCH_SIZE = 500  # Adjust batch size based on memory and DB performance
# -------------------

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)


# --- Helper Functions ---
async def get_column_info_async(sdp, table_name):
    query = f"""
        SELECT COLUMN_NAME, DATA_TYPE, CHARACTER_MAXIMUM_LENGTH
        FROM INFORMATION_SCHEMA.COLUMNS
        WHERE TABLE_NAME = '{table_name.split('.')[-1]}'
    """
    df = await sdp.fetch_data(query)
    col_info = {}
    for _, row in df.iterrows():
        col_info[row["COLUMN_NAME"].upper()] = {"type": row["DATA_TYPE"], "max_length": row["CHARACTER_MAXIMUM_LENGTH"]}
    return col_info


async def get_table_columns_async(sdp, table_name):
    query = f"""
        SELECT COLUMN_NAME
        FROM INFORMATION_SCHEMA.COLUMNS
        WHERE TABLE_NAME = '{table_name.split('.')[-1]}'
    """
    df = await sdp.fetch_data(query)
    return set(df["COLUMN_NAME"].str.upper())


# --- Main Processing Logic ---
async def main():
    logger.info("Initializing SDP and loading Excel workbook...")
    config = Config()
    # Use the optimized SDP class
    sdp = EnhancedSDP(config)

    # Load workbook and worksheet
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE_PATH)
        if EXCEL_TAB_NAME not in wb.sheetnames:
            logger.error(f"ERROR: Sheet '{EXCEL_TAB_NAME}' not found in workbook.")
            sys.exit(1)
        ws = wb[EXCEL_TAB_NAME]
    except FileNotFoundError:
        logger.error(f"ERROR: Excel file not found at '{EXCEL_FILE_PATH}'")
        sys.exit(1)
    except Exception as e:
        logger.error(f"Error loading Excel file: {e}")
        sys.exit(1)

    # Read and clean headers
    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    headers = [cell.value for cell in header_row]
    headers = clean_column_headers(headers)  # Assuming this returns uppercased headers

    pk_col_idx = -1  # Initialize column index

    # Check if the primary key field exists in the headers
    if PRIMARY_KEY_FIELD not in headers:
        logger.warning(f"Column '{PRIMARY_KEY_FIELD}' not found in Excel headers. Adding it to the right.")

        # Pass the name as a single-element list
        new_column_indices = add_columns_to_right(ws, [PRIMARY_KEY_FIELD])

        # Since we only added one column, get its index (the first element of the returned list)
        pk_col_idx = new_column_indices[0]

        # Also update the in-memory list of headers
        headers.append(PRIMARY_KEY_FIELD)  # Append because the function adds to the right

        logger.info(f"Added '{PRIMARY_KEY_FIELD}' as column {pk_col_idx} in the worksheet and updated headers list.")

    else:
        # If the column already exists, find its index (1-based)
        pk_col_idx = headers.index(PRIMARY_KEY_FIELD) + 1
        logger.info(f"Found existing '{PRIMARY_KEY_FIELD}' header at column index {pk_col_idx}.")

    # Add a check to ensure the index was determined successfully
    if pk_col_idx == -1:
        logger.error(f"Failed to determine the column index for '{PRIMARY_KEY_FIELD}'. Aborting.")
        sys.exit(1)

    # Build column name to index mapping (This part remains the same)
    colname_to_idx = {col: idx + 1 for idx, col in enumerate(headers) if col}

    # Build column name to index mapping (1-based for openpyxl)
    colname_to_idx = {col: idx + 1 for idx, col in enumerate(headers) if col}

    # --- Pre-fetch database info ---
    try:
        db_columns = await get_table_columns_async(sdp, DB_TABLE_NAME)
        col_info = await get_column_info_async(sdp, DB_TABLE_NAME)
        # Pre-fetch the table metadata for insertions
        _ = await sdp._get_table_object(DB_TABLE_NAME)  # Fetch and cache Table obj
    except Exception as e:
        logger.error(f"Failed to retrieve database metadata: {e}")
        sys.exit(1)
    # ------------------------------

    matching_cols = [col for col in headers if col and col.upper() in db_columns]
    logger.info("Matched columns between Excel and SQL table:")
    for col in matching_cols:
        logger.info(f" - {col}")

    # Ensure PK field itself is not treated as a column to insert if it exists in Excel
    cols_to_insert = [col for col in matching_cols if col != PRIMARY_KEY_FIELD]
    logger.info(f"\nColumns that will be used for INSERT: {', '.join(cols_to_insert)}")

    if PRIMARY_KEY_FIELD not in matching_cols and PRIMARY_KEY_FIELD in headers:
        logger.warning(
            f"Primary key field '{PRIMARY_KEY_FIELD}' exists in Excel but not matched in DB table '{DB_TABLE_NAME}'. It will not"
            " be inserted."
        )
    elif PRIMARY_KEY_FIELD not in headers:
        logger.info(f"Primary key field '{PRIMARY_KEY_FIELD}' will be generated by the database.")

    # Confirm
    proceed = input("\nProceed with this column mapping and insert? (y/n): ")
    if proceed.strip().lower() != "y":
        logger.info("Aborted by user.")
        sys.exit(0)

    # --- Batch Processing ---
    rows_to_insert_batch = []
    row_indices_batch = []  # To map results back to Excel rows
    processed_count = 0
    inserted_count = 0
    total_rows_in_sheet = ws.max_row - 1  # Exclude header

    logger.info(f"Starting data processing and insertion with batch size {BATCH_SIZE}...")

    async def process_batch():
        nonlocal inserted_count
        if not rows_to_insert_batch:
            return True  # Nothing to process

        batch_size = len(rows_to_insert_batch)
        first_row_idx = row_indices_batch[0] if row_indices_batch else "N/A"
        logger.info(f"Attempting to insert batch of {batch_size} rows (starting around Excel row {first_row_idx})...")

        try:
            # Use the batch insert method
            returned_pks = await sdp.insert_and_return_pks(DB_TABLE_NAME, rows_to_insert_batch, PRIMARY_KEY_FIELD)

            if returned_pks is None:
                # Error already logged in insert_many_and_return_pks
                logger.error(f"Batch insert failed for batch starting around row {first_row_idx}. Aborting.")
                return False  # Indicate failure

            if len(returned_pks) != len(row_indices_batch):
                logger.error(
                    f"CRITICAL: Mismatch in returned PK count ({len(returned_pks)}) and batch size ({len(row_indices_batch)}) for"
                    f" batch starting around row {first_row_idx}. Aborting."
                )
                # This suggests a major issue with the DB driver or logic.
                return False  # Indicate critical failure

            # Update Excel with returned PKs
            logger.info(f"Updating Excel in memory with {len(returned_pks)} new primary keys for the batch...")
            for pk_value, row_idx in zip(returned_pks, row_indices_batch):
                try:
                    # Ensure cell exists - openpyxl might create intermediate cells if needed
                    ws.cell(row=row_idx, column=pk_col_idx, value=pk_value)
                except Exception as cell_e:
                    logger.error(f"Failed to write PK {pk_value} to Excel cell (Row: {row_idx}, Col: {pk_col_idx}): {cell_e}")
                    return False

            inserted_count += len(returned_pks)

            logger.info(f"Saving Excel file '{EXCEL_FILE_PATH}' to persist {len(returned_pks)} newly added PKs...")
            try:
                wb.save(EXCEL_FILE_PATH)
                logger.info("Successfully saved Excel file after processing batch.")
            except Exception as save_e:
                logger.error(
                    f"CRITICAL: Failed to save Excel file after processing batch (starting row {first_row_idx}): {save_e}"
                )
                logger.error(
                    "Database inserts for this batch likely succeeded, but PKs could not be saved to Excel. Manual check"
                    " required."
                )
                return False  # Indicate failure

            # Clear the batch
            rows_to_insert_batch.clear()
            row_indices_batch.clear()
            logger.info(f"Batch processed and saved successfully. Total inserted so far: {inserted_count}")
            return True  # Indicate success

        except Exception as e:
            logger.error(
                f"An unexpected error occurred during database insertion for batch starting around row {first_row_idx}: {e}"
            )
            return False  # Indicate failure

    # --- Iterate through Excel rows ---
    try:
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            processed_count += 1
            pk_cell = row[pk_col_idx - 1]  # Use the determined pk_col_idx

            # Skip row if Primary Key cell already has a value
            if pk_cell.value is not None and str(pk_cell.value).strip() != "":
                continue

            # Build row dict only with columns that are matched and ARE NOT the PK field
            row_dict = {}
            for col_name in cols_to_insert:  # Use the pre-filtered list
                try:
                    cell_value = row[colname_to_idx[col_name] - 1].value
                    row_dict[col_name] = cell_value
                except KeyError:
                    logger.warning(
                        f"Column '{col_name}' not found in colname_to_idx map for row {row_idx}. Skipping column for this row."
                    )
                except IndexError:
                    logger.warning(
                        f"Index out of range for column '{col_name}' (expected index {colname_to_idx.get(col_name)}) in row"
                        f" {row_idx} (length {len(row)}). Skipping column."
                    )

            # Apply cleaning
            cleaned_row_dict = clean_row_values(row_dict, col_info)  # Pass db col info for cleaning context

            # Add data to batch
            rows_to_insert_batch.append(cleaned_row_dict)
            row_indices_batch.append(row_idx)

            # Process batch if full
            if len(rows_to_insert_batch) >= BATCH_SIZE:
                success = await process_batch()
                if not success:
                    raise RuntimeError("Batch processing failed. Aborting script.")  # Raise to break outer loop

            if processed_count % (BATCH_SIZE * 5) == 0:  # Log progress less frequently
                logger.info(f"Processed {processed_count}/{total_rows_in_sheet} rows...")

        # Process the final partial batch
        logger.info("Processing final batch...")
        success = await process_batch()
        if not success:
            raise RuntimeError("Final batch processing failed.")

    except Exception as e:
        logger.error(f"An error occurred during row iteration or batch processing: {e}")
        logger.error(f"Aborting after error. {inserted_count} rows may have been inserted and PKs potentially updated in Excel.")
        logger.info("Attempting to save Excel file with potentially partial updates...")
        try:
            wb.save(EXCEL_FILE_PATH)
            logger.info(f"Excel file saved: {EXCEL_FILE_PATH}")
        except Exception as save_e:
            logger.error(f"CRITICAL: Failed to save Excel file after error: {save_e}")
        sys.exit(1)

    # --- Finalization ---
    logger.info("Saving final updates to Excel file...")
    try:
        wb.save(EXCEL_FILE_PATH)
        logger.info(f"Excel file saved successfully: {EXCEL_FILE_PATH}")
        logger.info(f"Script finished. Total rows processed: {processed_count}. Total new rows inserted: {inserted_count}.")
        logger.info(f"Primary keys written to Excel in column '{PRIMARY_KEY_FIELD}'.")
    except Exception as e:
        logger.error(f"Failed to save the final Excel file: {e}")
        logger.error("Database insertions likely completed, but PKs could not be saved back to Excel.")
        sys.exit(1)


if __name__ == "__main__":
    asyncio.run(main())
