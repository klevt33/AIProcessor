# step3_results_to_excel.py

import asyncio
import logging
import os
import sys
from datetime import datetime
from typing import Dict, List, Optional

import openpyxl
import pandas as pd
from config_e2e import EXCEL_FILE_PATH, EXCEL_TAB_NAME, PRIMARY_KEY_FIELD, add_columns_to_right, fetch_data_in_batches

from config import Config
from constants import DatabaseObjects
from matching_utils import read_manufacturer_data
from sdp import SDP
from utils import get_current_datetime_cst

# Configuration
PREFIX = "out_"
ID_FIELDS_TO_EXCLUDE = ["IVCE_DTL_UID", "IVCE_DTL_ID", "IVCE_TRKG_MSTR_UID", "IVCE_HDR_ID"]
BATCH_SIZE = 500  # SQL batch size
BACKUP_BEFORE_SAVE = True

# Set up logging
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)


def backup_excel_file(file_path: str) -> Optional[str]:
    """Create a backup of the Excel file before modifying it"""
    if not BACKUP_BEFORE_SAVE:
        return None

    try:
        backup_path = f"{file_path}.{get_current_datetime_cst()}.bak"
        if os.path.exists(file_path):
            import shutil

            shutil.copy2(file_path, backup_path)
            logger.info(f"Created backup at {backup_path}")
            return backup_path
    except Exception as e:
        logger.warning(f"Failed to create backup: {e}")

    return None


async def get_database_data(sdp: SDP, excel_pks: List[str]) -> pd.DataFrame:
    """Fetch data from database for the given primary keys"""
    logger.info(f"Fetching data for {len(excel_pks)} primary keys from database...")

    dtl_tbl = DatabaseObjects.TBL_INVC_DTL
    trkg_tbl = DatabaseObjects.TBL_IVCE_TRKG_MSTR
    dtl_pk = "IVCE_DTL_UID"
    trkg_pk = "IVCE_DTL_ID"  # Assuming this is the FK in trkg_tbl matching dtl_pk

    # Original query structure
    base_query = f"""
    SELECT
        a.*,
        b.*
    FROM {dtl_tbl} a
    LEFT JOIN {trkg_tbl} b ON a.{dtl_pk} = b.{trkg_pk}
    WHERE a.{dtl_pk} IN ({{pk_list}})
    """

    try:
        # Original fetch call
        df = await fetch_data_in_batches(sdp, base_query, excel_pks, batch_size=BATCH_SIZE)
        logger.info(f"Retrieved {len(df)} rows from database")
        return df
    except Exception as e:
        logger.error(f"Database error: {e}")
        raise


def process_dataframe(df: pd.DataFrame, manufacturer_mapping: dict) -> pd.DataFrame:
    """Process the dataframe to prepare it for Excel export"""
    if df.empty:
        # Return empty dataframe as per original logic if input is empty
        return df

    logger.info("Processing database results...")
    df_processed = df.copy()  # Work on a copy to avoid modifying the original passed df

    # --- Add Clean Manufacturer Name ---
    clean_mfr_col = "clean_MFR_NM"
    raw_mfr_col = "MFR_NM"  # "MFR_NM"
    out_clean_mfr_col = f"{PREFIX}{clean_mfr_col}"  # Target output name

    if raw_mfr_col in df_processed.columns:
        logger.info(f"Generating clean manufacturer names using mapping for column '{raw_mfr_col}'")
        # Apply mapping using .get for safety (returns '' if key not found)
        df_processed[clean_mfr_col] = df_processed[raw_mfr_col].apply(
            lambda x: manufacturer_mapping.get(str(x).strip().upper(), "") if pd.notna(x) else ""
        )
        logger.info(f"Created '{clean_mfr_col}' column.")
        created_clean_mfr = True
    else:
        logger.warning(f"Column '{raw_mfr_col}' not found in database results. Cannot create clean manufacturer column.")
        created_clean_mfr = False

    # Remove ID columns (keep PK temporarily)
    # Ensure PRIMARY_KEY_FIELD exists before trying to exclude it from drop list
    cols_to_drop = []
    if PRIMARY_KEY_FIELD in df_processed.columns:
        cols_to_drop = [col for col in ID_FIELDS_TO_EXCLUDE if col in df_processed.columns and col != PRIMARY_KEY_FIELD]
    else:
        # If PK is not even in the df, just drop whatever from ID_FIELDS_TO_EXCLUDE is present
        logger.warning(f"Primary key '{PRIMARY_KEY_FIELD}' not found in DataFrame columns for exclusion check.")
        cols_to_drop = [col for col in ID_FIELDS_TO_EXCLUDE if col in df_processed.columns]

    if cols_to_drop:
        df_processed = df_processed.drop(columns=cols_to_drop)

    # Add prefix to all *remaining* columns for Excel output
    # Need to exclude the primary key field from prefixing if it's still present
    cols_to_prefix = [col for col in df_processed.columns if col != PRIMARY_KEY_FIELD]
    col_rename_map = {col: f"{PREFIX}{col}" for col in cols_to_prefix}
    df_processed = df_processed.rename(columns=col_rename_map)

    # Get all columns starting with the prefix
    current_out_cols = [col for col in df_processed.columns if col.startswith(PREFIX)]
    ordered_out_cols = []
    # Place the clean manufacturer column first if it exists among prefixed columns
    if created_clean_mfr and out_clean_mfr_col in current_out_cols:
        ordered_out_cols.append(out_clean_mfr_col)
        ordered_out_cols.extend([col for col in current_out_cols if col != out_clean_mfr_col])
    else:
        # If not created or not prefixed, use the current order
        ordered_out_cols = current_out_cols

    # Reconstruct the final column list including the PK (if present) and the ordered output columns
    final_columns_order = []
    if PRIMARY_KEY_FIELD in df_processed.columns:  # Check if PK survived drops/renames
        final_columns_order.append(PRIMARY_KEY_FIELD)
    final_columns_order.extend(ordered_out_cols)

    # Reindex, ensuring columns exist in df_processed
    final_columns_order = [col for col in final_columns_order if col in df_processed.columns]
    df_processed = df_processed[final_columns_order]
    # --- End New Column Ordering Logic ---

    # Drop columns where all values are NaN or empty
    df_processed = df_processed.dropna(axis=1, how="all")

    # Remove duplicate columns (keep first)
    df_processed = df_processed.loc[:, ~df_processed.columns.duplicated(keep="first")]

    logger.info(f"Processed dataframe has {len(df_processed.columns)} columns after cleanup")
    return df_processed  # Return the fully processed dataframe


def get_excel_primary_keys(ws: openpyxl.worksheet.worksheet.Worksheet) -> Dict[str, int]:
    """Extract primary keys and their row indices from Excel"""
    logger.info("Reading primary keys from Excel...")

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    if PRIMARY_KEY_FIELD not in headers:
        raise ValueError(f"Required column '{PRIMARY_KEY_FIELD}' not found in Excel")

    pk_col_idx = headers.index(PRIMARY_KEY_FIELD)  # Original used 0-based index for slicing below

    excel_pks = {
        str(row[pk_col_idx].value): idx
        for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2)
        if row[pk_col_idx].value is not None and str(row[pk_col_idx].value).strip() != ""
    }

    logger.info(f"Found {len(excel_pks)} populated primary keys in Excel")
    return excel_pks


def safe_write_value(cell, value):
    """Safely write a value to an Excel cell with appropriate type handling"""
    try:
        # Original logic structure
        cell.value = None if pd.isna(value) else (value if isinstance(value, (int, float, str, datetime)) else str(value))
    except Exception as e:
        logger.warning(f"Error writing value {value} to cell: {e}")
        # Original fallback
        cell.value = str(value)


async def main():
    logger.info(f"Starting Excel update process for {EXCEL_FILE_PATH}")

    # Keep original variable names and flow
    config = None
    sdp = None
    wb = None  # Initialize wb

    try:
        # Initialize database connection
        config = Config()
        sdp = SDP(config)

        logger.info("Fetching manufacturer mapping...")
        manufacturer_mapping = await read_manufacturer_data(sdp)
        logger.info(f"Loaded {len(manufacturer_mapping)} manufacturer mappings.")

        # Create backup
        backup_excel_file(EXCEL_FILE_PATH)

        try:
            # Load Excel
            logger.info(f"Loading Excel workbook: {EXCEL_FILE_PATH}")
            wb = openpyxl.load_workbook(EXCEL_FILE_PATH)

            if EXCEL_TAB_NAME not in wb.sheetnames:
                logger.error(f"Sheet '{EXCEL_TAB_NAME}' not found in workbook.")
                return 1  # Original exit path

            ws = wb[EXCEL_TAB_NAME]

            # Extract primary keys from Excel
            row_idx_map = get_excel_primary_keys(ws)

            if not row_idx_map:
                logger.warning("No populated PKs found in Excel. Nothing to update.")
                return 0  # Original exit path

            # Query database
            excel_pks = list(row_idx_map.keys())
            # Changed variable name from df to db_data_raw for clarity before processing
            db_data_raw = await get_database_data(sdp, excel_pks)

            if db_data_raw.empty:
                logger.warning("No matching records found in database for Excel PKs.")

            # Process database results, passing the mapping
            df_processed = process_dataframe(db_data_raw, manufacturer_mapping)

            # Map primary key to processed database results
            # Ensure PK column is string type for consistent matching with excel_pks keys
            df_processed[PRIMARY_KEY_FIELD] = df_processed[PRIMARY_KEY_FIELD].astype(str)

            # Create the lookup dictionary using the processed data
            sql_data = {
                row[PRIMARY_KEY_FIELD]: row
                for _, row in df_processed.iterrows()
                if row[PRIMARY_KEY_FIELD] in row_idx_map  # Only map rows that have a corresponding PK in Excel
            }

            common_keys = set(row_idx_map.keys()).intersection(set(sql_data.keys()))
            logger.info(
                f"Keys in Excel: {len(row_idx_map)}, Keys mapped from DB: {len(sql_data)}, Matching keys for update:"
                f" {len(common_keys)}"
            )

            # Get list of new columns for Excel (those starting with PREFIX)
            new_cols_out = [col for col in df_processed.columns if col.startswith(PREFIX)]
            logger.info(f"Adding/updating {len(new_cols_out)} columns in Excel: {new_cols_out}")

            if not new_cols_out:
                logger.info("No columns with prefix found after processing. No columns to add/update.")
            else:
                # Add new columns to Excel header
                col_indices = add_columns_to_right(ws, new_cols_out)
                # Create map from prefixed column name to Excel column index (1-based)
                colname_to_excel_idx = dict(zip(new_cols_out, col_indices))

                # Write database results into Excel
                logger.info("Writing database results to Excel...")
                updated_count = 0

                # Iterate through the primary keys found *in Excel*
                for pk_str, row_idx in row_idx_map.items():
                    # Check if this Excel PK has corresponding processed data
                    if pk_str in sql_data:
                        row_data_series = sql_data[pk_str]  # Get the processed data Series

                        # Iterate through the output columns to be written
                        for out_col in new_cols_out:
                            # Get value from the processed data Series using the prefixed column name
                            # Use .get(out_col, None) for safety if a column is somehow missing from a specific row Series
                            value = row_data_series.get(out_col)

                            # Original write logic using the helper function
                            safe_write_value(ws.cell(row=row_idx, column=colname_to_excel_idx[out_col]), value)

                        updated_count += 1
                    # else: PK from Excel not found in DB results, do nothing for this row (original behavior)

                # Original final log message structure
                logger.info(f"Saving updated Excel workbook to {EXCEL_FILE_PATH}")
                wb.save(EXCEL_FILE_PATH)
                logger.info(f"Successfully updated {updated_count} Excel rows with data for {len(new_cols_out)} columns.")

        finally:
            # Original cleanup logic
            if wb:  # Check if wb was successfully assigned
                wb.close()

        return 0  # Original success return

    except Exception as e:
        # Original generic error handling
        logger.error(f"Error in main process: {e}", exc_info=True)
        return 1  # Original failure return
    finally:
        # Ensure SDP connection is closed if it was opened
        if sdp and hasattr(sdp, "close"):
            try:
                await sdp.close()
                logger.info("Database connection closed.")
            except Exception as e_close:
                logger.warning(f"Error closing database connection: {e_close}")


if __name__ == "__main__":
    # Original __main__ block
    exit_code = asyncio.run(main())
    sys.exit(exit_code)
