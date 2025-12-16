"""
Shared Utilities Module
"""

import logging
import math
import os
import sys
from typing import Any, Dict, List, Optional, Union

import openpyxl  # For ExcelUtil
from sqlalchemy import MetaData, Table, insert

file_path = os.path.abspath(__file__)
root = os.path.abspath(os.path.join(file_path, "..", ".."))
if os.path.isdir(root) and root not in sys.path:
    sys.path.insert(0, root)
    APP_ROOT = root
    print(f"[INFO] Added to sys.path: {root}")

from constants import DatabaseObjects  # For DatabaseObjects.TBL_INVC_DTL
from sdp import SDP  # Base class for EnhancedSDP

# Configure logging
logger = logging.getLogger(__name__)  # Logger for this module


def setup_logging(level=logging.INFO):
    """Set up logging with appropriate formatting."""
    # Check if root logger already has handlers to avoid duplicate logs if main app configures it
    root_logger = logging.getLogger()
    if not root_logger.hasHandlers() or not root_logger.handlers:
        logging.basicConfig(level=level, format="%(asctime)s - %(levelname)s - %(name)s - %(message)s")
    else:
        # If already configured, just set the level for this specific logger
        logger.setLevel(level)


class Config:
    """Configuration settings for the module."""

    # EXCEL = {
    #     "FILE_PATH": os.getenv(
    #         "E2E_EXCEL_FILE_PATH", r"C:\Users\LevtovKirill(Perfici\OneDrive - ArchKey Solutions\Documents\ProdRun\Prod-Test.xlsx"
    #     ),
    #     "TAB_NAME": os.getenv("E2E_EXCEL_TAB_NAME", "Sheet1"),
    #     "PRIMARY_KEY_FIELD": "IVCE_DTL_UID",
    #     "REQUEST_ID_FIELD": "REQUEST_ID",
    # }

    EXCEL = {
        "FILE_PATH": os.getenv("E2E_EXCEL_FILE_PATH", r"C:\Users\LevtovKirill(Perfici\Downloads\Test-AI-300.xlsx"),
        "TAB_NAME": os.getenv("E2E_EXCEL_TAB_NAME", "invoice_test_data"),
        "PRIMARY_KEY_FIELD": "IVCE_DTL_UID",
        "REQUEST_ID_FIELD": "REQUEST_ID",
    }

    DATABASE = {"TABLE_NAME": DatabaseObjects.TBL_INVC_DTL, "BATCH_SIZE": 500}

    @classmethod
    def get_excel_path(cls) -> str:
        """Get the Excel file path with validation."""
        if not os.path.exists(cls.EXCEL["FILE_PATH"]):
            raise FileNotFoundError(f"Excel file not found: {cls.EXCEL['FILE_PATH']}")
        return cls.EXCEL["FILE_PATH"]


class EnhancedSDP(SDP):
    """
    A subclass of SDP that provides enhanced functionality for database operations,
    particularly for insert operations that return primary keys.
    """

    def __init__(self, config):  # Expects the main application config object
        # Call the parent class constructor
        super().__init__(config)
        # Initialize the table cache
        self._table_cache = {}  # Cache for SQLAlchemy Table objects

    async def _get_table_object(self, table_name: str) -> Table:
        """Gets or creates the SQLAlchemy Table object, caching it."""
        if table_name in self._table_cache:
            return self._table_cache[table_name]

        logger.info(f"Fetching metadata for table: {table_name}")
        if "." in table_name:
            schema, tbl = table_name.split(".")
        else:
            schema, tbl = None, table_name

        metadata = MetaData()
        try:
            # Use synchronous reflection as the base SDP uses sync engine
            # If you upgrade to async engine, this needs await
            # Ensure self.sqlalchemy_engine is available from base SDP class
            if not hasattr(self, "sqlalchemy_engine") or self.sqlalchemy_engine is None:
                raise AttributeError(
                    "EnhancedSDP requires 'sqlalchemy_engine' from the base SDP, initialized with main app config."
                )
            table = Table(tbl, metadata, schema=schema, autoload_with=self.sqlalchemy_engine)
            self._table_cache[table_name] = table
            logger.info(f"Successfully fetched metadata for {table_name}")
            return table
        except Exception as e:
            logger.error(f"Failed to reflect table {table_name}: {e}", exc_info=True)
            raise  # Re-raise the exception to stop the process if metadata fails

    async def insert_and_return_pks(self, table_name: str, values_dicts: list[dict], pk_field: str):
        """
        Inserts multiple records into the specified table using executemany
        and returns the corresponding primary key values in order.
        (This is the original working version from your provided file)
        """
        if not values_dicts:
            return []

        try:
            table = await self._get_table_object(table_name)

            if pk_field not in table.c:
                raise ValueError(f"Primary key field '{pk_field}' not found in table '{table_name}' metadata.")

            stmt = insert(table).returning(table.c[pk_field])

            pks = []
            # Use a synchronous context since the base SDP uses a sync engine (as per original comments)
            with self.sqlalchemy_engine.begin() as conn:  # This was the line causing issues if it was 'async with'
                result = conn.execute(stmt, values_dicts)
                pks = [row[0] for row in result.fetchall()]

            if len(pks) != len(values_dicts):
                logger.error(
                    f"Mismatch between inserted rows ({len(values_dicts)}) and returned PKs ({len(pks)}) for table {table_name}."
                )
                raise RuntimeError("Number of returned PKs does not match number of inserted rows.")

            return pks
        except Exception as e:
            logger.error(f"Error in insert_and_return_pks for table {table_name}: {e}", exc_info=True)
            if values_dicts:
                logger.error(f"First row data in failed batch: {values_dicts[0]}")
            return None


class ExcelUtil:
    """Excel related utility functions."""

    @staticmethod
    def get_ids_from_excel(
        file_path: str, sheet_name: str, pk_col: str  # pk_col is the header name of the primary key column
    ) -> List[str]:
        """Original get_ids_from_excel method."""
        wb = None
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found in workbook '{file_path}'.")

            ws = wb[sheet_name]

            headers = None
            col_idx = None  # 0-based index

            # First pass - just get headers and find index
            # ws.iter_rows(min_row=1, max_row=1, values_only=True) yields a tuple of rows,
            # and each row is a tuple of cell values. We need the first (and only) row.
            header_row_values = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if header_row_values is None:
                raise ValueError(f"Header row not found in sheet '{sheet_name}'.")

            headers = header_row_values  # headers is now a tuple of actual header values

            if pk_col in headers:
                col_idx = headers.index(pk_col)

            if col_idx is None:
                raise ValueError(f"Column '{pk_col}' not found in Excel headers: {headers}")

            ids = []
            for row_values in ws.iter_rows(min_row=2, values_only=True):  # Start from data rows
                if col_idx < len(row_values):  # Check if row has enough cells
                    pk_val = row_values[col_idx]
                    if pk_val is not None and str(pk_val).strip() != "":
                        ids.append(str(pk_val))
            return ids
        finally:
            if wb:
                wb.close()

    @staticmethod
    def get_unprocessed_ids_from_excel(file_path: str, sheet_name: str, pk_col_name: str, request_id_col_name: str) -> List[str]:
        """
        Reads an Excel file and returns a list of Primary Key IDs ONLY for rows
        where the Request ID column is empty. This enables resuming a failed run.
        """
        wb = None
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found in workbook.")
            ws = wb[sheet_name]

            # --- THIS IS THE CORRECTED LINE ---
            # It now correctly converts the tuple of header values into a list.
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not header_row:
                raise ValueError(f"Header row not found in sheet '{sheet_name}'.")
            headers = list(header_row)
            # --- END OF CORRECTION ---

            try:
                pk_col_idx = headers.index(pk_col_name)
            except ValueError:
                raise ValueError(f"Primary Key column '{pk_col_name}' not found in Excel: {headers}")

            request_id_col_idx = -1
            if request_id_col_name in headers:
                request_id_col_idx = headers.index(request_id_col_name)
            else:
                # If the REQUEST_ID column doesn't exist yet, it's the very first run.
                # In this case, all IDs are considered unprocessed.
                logger.info(f"Column '{request_id_col_name}' not found. Assuming first run, processing all IDs.")
                all_ids = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    # Check if the row has enough columns to contain the PK
                    if len(row) > pk_col_idx:
                        pk_val = row[pk_col_idx]
                        if pk_val is not None and str(pk_val).strip() != "":
                            all_ids.append(str(pk_val))
                return all_ids

            # If the REQUEST_ID column exists, find only the rows where it's empty.
            unprocessed_ids = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                # Check if the row is long enough to have both columns
                if len(row) > max(pk_col_idx, request_id_col_idx):
                    request_id_val = row[request_id_col_idx]
                    if request_id_val is None or str(request_id_val).strip() == "":
                        pk_val = row[pk_col_idx]
                        if pk_val is not None and str(pk_val).strip() != "":
                            unprocessed_ids.append(str(pk_val))

            return unprocessed_ids
        finally:
            if wb:
                wb.close()

    @staticmethod
    def clean_column_headers(header_row_values: List[Any]) -> List[Optional[str]]:  # Accepts list of values
        """Original clean_column_headers method."""
        seen = set()
        cleaned = []
        for col_val in header_row_values:  # Iterate over values directly
            # Original logic: col_clean = col.strip().upper() if isinstance(col, str) else col
            # Assuming col_val can be non-string (e.g. number from header), convert to str first
            if col_val is None:
                col_str = ""
            else:
                col_str = str(col_val)

            col_clean = col_str.strip().upper() if col_str else None  # Handle if col_str is empty after str(None)

            if col_clean and col_clean not in seen:  # Also filter empty headers if col_clean is None or ""
                cleaned.append(col_clean)
                seen.add(col_clean)
            elif col_clean and col_clean in seen:  # If it's a duplicate
                logger.warning(f"Duplicate cleaned header '{col_clean}' found (original: '{col_val}'). Will be None.")
                cleaned.append(None)
            else:  # If col_clean is None or "" (e.g. empty original header cell)
                cleaned.append(None)
        return cleaned

    @staticmethod
    def add_columns_to_right(ws: openpyxl.worksheet.worksheet.Worksheet, col_names: List[str]) -> List[int]:
        """Original add_columns_to_right method."""
        max_col = ws.max_column
        new_col_indices = []  # Store 1-based indices
        for i, col_name in enumerate(col_names):
            current_col_idx = max_col + i + 1
            ws.cell(row=1, column=current_col_idx, value=col_name)
            new_col_indices.append(current_col_idx)
        return new_col_indices

    # --- ADDED/ENHANCED METHODS FOR COSMOS SCRIPT ---
    @staticmethod
    def update_excel_with_request_id(
        file_path: str,
        sheet_name: str,
        primary_key_column_name: str,  # PK column in Excel
        ids_in_batch: List[str],  # List of PK values for rows to update
        request_id_value: str,
        request_id_header_name: str,  # Name of the Request ID column in Excel
    ):
        """Writes a request_id to specific Excel rows. Adds column if needed. Saves file."""
        wb = None
        try:
            wb = openpyxl.load_workbook(file_path)
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found in workbook '{file_path}'.")
            sheet = wb[sheet_name]
            header_cells = sheet[1]
            header_values = [str(cell.value).strip() for cell in header_cells if cell.value is not None]

            request_id_col_num = -1
            try:
                request_id_col_idx_0_based = header_values.index(request_id_header_name)
                request_id_col_num = request_id_col_idx_0_based + 1
            except ValueError:
                request_id_col_num = sheet.max_column + 1
                sheet.cell(row=1, column=request_id_col_num, value=request_id_header_name)
                logger.info(f"Added new column '{request_id_header_name}' at col {request_id_col_num} in '{sheet_name}'.")

            pk_col_num = -1
            try:
                pk_col_idx_0_based = header_values.index(primary_key_column_name)
                pk_col_num = pk_col_idx_0_based + 1
            except ValueError:
                raise ValueError(
                    f"PK column '{primary_key_column_name}' not found in sheet '{sheet_name}'. Headers: {header_values}"
                )

            ids_in_batch_set = set(map(str, ids_in_batch))
            updated_rows_count = 0
            for row_idx in range(2, sheet.max_row + 1):
                pk_cell_value = sheet.cell(row=row_idx, column=pk_col_num).value
                if pk_cell_value is not None and str(pk_cell_value).strip() in ids_in_batch_set:
                    sheet.cell(row=row_idx, column=request_id_col_num, value=request_id_value)
                    updated_rows_count += 1

            if updated_rows_count > 0:
                logger.info(f"Updated {updated_rows_count} rows in '{sheet_name}' with RequestID '{request_id_value}'.")

            wb.save(file_path)
        finally:
            if wb:
                wb.close()

    @staticmethod
    def get_row_identifiers_for_cosmos_matching(
        file_path: str,
        sheet_name: str,
        pk_excel_header: str,  # Header name for PK (e.g., IVCE_DTL_UID)
        request_id_excel_header: Optional[str],  # Header name for Request ID (e.g., REQUEST_ID)
    ) -> List[Dict[str, Any]]:
        """Reads Excel for PK and optional Request ID for matching with CosmosDB."""
        identifiers = []
        wb = None
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found in workbook '{file_path}'.")
            ws = wb[sheet_name]

            header_row_vals = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not header_row_vals:
                raise ValueError(f"No header row in sheet '{sheet_name}'.")

            try:
                pk_col_idx = list(header_row_vals).index(pk_excel_header)  # Use original tuple for index
            except ValueError:
                raise ValueError(f"PK column '{pk_excel_header}' not found in sheet '{sheet_name}'. Headers: {header_row_vals}")

            request_id_col_idx = None
            if request_id_excel_header:
                try:
                    request_id_col_idx = list(header_row_vals).index(request_id_excel_header)
                except ValueError:
                    logger.info(
                        f"Request ID column '{request_id_excel_header}' not found in sheet '{sheet_name}'. Proceeding without it"
                        " for matching."
                    )

            for rn, row_cells_tuple in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                pk_val_str = None
                if pk_col_idx < len(row_cells_tuple):
                    cell_val = row_cells_tuple[pk_col_idx]
                    if cell_val is not None:
                        val_str = str(cell_val).strip()
                        if val_str:
                            pk_val_str = val_str

                if not pk_val_str:
                    continue  # Skip row if PK is empty

                request_id_val_str = None
                if request_id_col_idx is not None and request_id_col_idx < len(row_cells_tuple):
                    cell_val = row_cells_tuple[request_id_col_idx]
                    if cell_val is not None:
                        val_str = str(cell_val).strip()
                        if val_str:
                            request_id_val_str = val_str

                identifiers.append(
                    {
                        "excel_row_num": rn,  # 1-based Excel row number
                        "pk_value": pk_val_str,
                        "request_id_value": request_id_val_str,
                    }
                )
            return identifiers
        finally:
            if wb:
                wb.close()

    @staticmethod
    def get_ids_for_reprocessing(
        file_path: str,
        sheet_name: str,
        pk_col_name: str,
        reprocessing_trigger_columns: List[str],  # e.g., ["REQUEST_ID", "process_details_1", ...]
        error_check_columns: List[str],  # e.g., ["process_details_1", ..., "process_details_5"]
        error_substrings: List[str],
    ) -> Optional[List[str]]:
        """
        Checks if Excel is set up for reprocessing and returns IDs of rows with errors.
        Returns None if reprocessing mode should not be triggered (e.g., trigger columns missing).
        Returns a list of PKs if reprocessing mode is active (can be empty if no errors found).
        """
        wb = None
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)  # data_only=True to get values
            if sheet_name not in wb.sheetnames:
                logger.warning(f"Sheet '{sheet_name}' not found in '{file_path}' for reprocessing check.")
                return None  # Cannot determine mode

            ws = wb[sheet_name]

            header_row_values = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not header_row_values:
                logger.warning(f"No header row in sheet '{sheet_name}' for reprocessing check.")
                return None

            # Convert header tuple to list of strings for easier checking
            headers = [str(h).strip() for h in header_row_values if h is not None]

            # 1. Check if all reprocessing_trigger_columns exist
            for col_name in reprocessing_trigger_columns:
                if col_name not in headers:
                    logger.info(f"Reprocessing trigger column '{col_name}' not found. Defaulting to full processing mode.")
                    return None  # Not in reprocessing mode

            logger.info("All reprocessing trigger columns found. Checking for errors to determine IDs for reprocessing.")

            # 2. Get column indices for PK and error check columns
            try:
                pk_col_idx = headers.index(pk_col_name)
            except ValueError:
                logger.error(f"Primary key column '{pk_col_name}' not found, though reprocessing columns exist. Cannot proceed.")
                # This is an inconsistent state, but for mode determination, we'd still not reprocess.
                # Or raise an error if this state is invalid. For now, treat as "not reprocessing specific errors."
                return None

            error_col_indices: Dict[str, int] = {}
            for err_col_name in error_check_columns:
                try:
                    error_col_indices[err_col_name] = headers.index(err_col_name)
                except ValueError:
                    # This should not happen if reprocessing_trigger_columns check passed and they are part of error_check_columns
                    logger.warning(
                        f"Error check column '{err_col_name}' not found in headers. It will be skipped for error checking."
                    )
                    # Continue, but this column won't be checked.

            ids_with_errors: List[str] = []
            for row_num, row_values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                pk_value = None
                if pk_col_idx < len(row_values):
                    pk_cell_val = row_values[pk_col_idx]
                    if pk_cell_val is not None and str(pk_cell_val).strip() != "":
                        pk_value = str(pk_cell_val).strip()

                if not pk_value:  # Skip rows without a PK
                    continue

                row_has_error = False
                for err_col_name, err_col_idx in error_col_indices.items():
                    if err_col_idx < len(row_values):
                        cell_content = row_values[err_col_idx]
                        if cell_content is not None:
                            cell_content_str = str(cell_content)  # Ensure it's a string
                            for error_str in error_substrings:
                                if error_str in cell_content_str:
                                    logger.debug(
                                        f"Error substring '{error_str}' found in row {row_num}, column '{err_col_name}' for PK"
                                        f" '{pk_value}'."
                                    )
                                    row_has_error = True
                                    break  # Found an error string in this cell
                    if row_has_error:
                        break  # Found an error in one of the columns for this row

                if row_has_error:
                    ids_with_errors.append(pk_value)

            logger.info(f"Found {len(ids_with_errors)} IDs with errors for reprocessing.")
            return ids_with_errors

        except Exception as e:
            logger.error(f"Error during get_ids_for_reprocessing: {e}", exc_info=True)
            return None  # Fallback to normal mode on error
        finally:
            if wb:
                wb.close()


class DataUtil:
    """Data cleaning and processing utilities."""

    @staticmethod
    def clean_row_values(row_dict: Dict[str, Any], col_info: Dict[str, Dict[str, Any]]) -> Dict[str, Any]:
        """Original clean_row_values method."""
        cleaned = {}
        for col, v in row_dict.items():
            # col_info keys are expected to be uppercase from get_column_info_async
            col_upper_for_lookup = col.upper() if isinstance(col, str) else str(col).upper()
            info = col_info.get(col_upper_for_lookup, {})
            max_length = info.get("max_length")

            if v is None or (isinstance(v, str) and (v.strip().upper() == "NULL" or v.strip() == "")):
                cleaned[col] = None
            elif isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
                cleaned[col] = None
            elif (
                isinstance(v, str)
                and max_length is not None
                and not (isinstance(max_length, float) and math.isnan(max_length))  # Check for NaN max_length
                and len(v) > int(max_length)
            ):
                logger.warning(f"Value for column '{col}' (len: {len(v)}) > max_length ({int(max_length)}). Truncating.")
                cleaned[col] = v[: int(max_length)]
            else:
                cleaned[col] = v
        return cleaned


class DatabaseUtil:
    """Database utility functions."""

    @staticmethod
    async def fetch_data_in_batches(
        sdp: SDP,  # Expects an SDP-like object
        base_query: str,
        pk_list: List[Union[str, int]],
        batch_size: int = 500,  # Original default
        show_progress: bool = False,
    ):
        """Original fetch_data_in_batches method."""
        import pandas as pd  # Local import

        if not pk_list:
            logger.warning("Empty primary key list provided to fetch_data_in_batches")
            return pd.DataFrame()

        all_results = []
        num_total_batches = (len(pk_list) + batch_size - 1) // batch_size

        current_iterator = range(0, len(pk_list), batch_size)
        if show_progress:
            try:
                # Assuming sdp.fetch_data is async, use tqdm.asyncio if available
                # If sdp.fetch_data is synchronous but called from async code,
                # the progress bar might not update ideally without `asyncio.to_thread`.
                # For simplicity, using standard tqdm and assuming it's okay for E2E script context.
                from tqdm import tqdm

                current_iterator = tqdm(current_iterator, total=num_total_batches, desc="Fetching data in batches")
            except ImportError:
                logger.warning("tqdm not installed, progress bar disabled for fetch_data_in_batches.")

        for i_batch, start_index in enumerate(current_iterator):
            batch_pks = pk_list[start_index : start_index + batch_size]

            if not batch_pks:
                continue

            # Original logic for formatting pk_str for IN clause
            # Ensure consistent type checking (first element determines type for whole batch)
            first_pk_is_numeric = isinstance(batch_pks[0], (int, float)) and not isinstance(batch_pks[0], bool)

            if first_pk_is_numeric:
                # All PKs in batch assumed numeric, no quotes
                pk_values_str = ",".join(map(str, batch_pks))
            else:
                # All PKs in batch assumed string, add quotes and escape internal quotes
                pk_values_str = ",".join(f"'{str(pk).replace("'", "''")}'" for pk in batch_pks)

            try:
                query_with_pks = base_query.format(pk_list=pk_values_str)
                # logger.debug(f"Executing batch query {i_batch+1}/{num_total_batches}: {query_with_pks[:250]}...")
                df_batch = await sdp.fetch_data(query_with_pks)  # Assumes sdp.fetch_data is async
                if not df_batch.empty:
                    all_results.append(df_batch)
            except Exception as e:
                logger.error(
                    f"Error fetching data for batch {i_batch + 1} (Query: {base_query[:100]}... PKs like:"
                    f" {batch_pks[0] if batch_pks else 'N/A'}): {e}",
                    exc_info=True,
                )
                # Depending on requirements, you might re-raise or continue
                # For now, let's try to process other batches:
                # raise # Uncomment to stop on first batch error

        if all_results:
            return pd.concat(all_results, ignore_index=True)
        else:
            logger.info("No data found for the provided primary keys after processing all batches.")
            return pd.DataFrame()


# --- Backward compatibility aliases (original set) ---
get_ids_from_excel = ExcelUtil.get_ids_from_excel
clean_column_headers = ExcelUtil.clean_column_headers  # This was the problematic one due to input type
clean_row_values = DataUtil.clean_row_values
add_columns_to_right = ExcelUtil.add_columns_to_right
fetch_data_in_batches = DatabaseUtil.fetch_data_in_batches

# --- Module-level Constants for backward compatibility (original set) ---
PRIMARY_KEY_FIELD = Config.EXCEL["PRIMARY_KEY_FIELD"]
EXCEL_FILE_PATH = Config.EXCEL["FILE_PATH"]
EXCEL_TAB_NAME = Config.EXCEL["TAB_NAME"]
DB_TABLE_NAME = Config.DATABASE["TABLE_NAME"]

# --- ADDED Aliases for new ExcelUtil methods ---
update_excel_with_request_id = ExcelUtil.update_excel_with_request_id
get_row_identifiers_for_cosmos_matching = ExcelUtil.get_row_identifiers_for_cosmos_matching
get_ids_for_reprocessing = ExcelUtil.get_ids_for_reprocessing
get_unprocessed_ids_from_excel = ExcelUtil.get_unprocessed_ids_from_excel

# --- ADDED Constant for new EXCEL field ---
# This is already part of Config.EXCEL, but if scripts import it directly:
REQUEST_ID_FIELD = Config.EXCEL["REQUEST_ID_FIELD"]


# Initialize logging when module is imported
setup_logging()
