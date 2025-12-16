import asyncio
import json
import logging
import sys
from typing import Any, Dict, List, Optional

import openpyxl

# --- E2E Test Suite specific imports ---
from config_e2e import (
    EXCEL_FILE_PATH,
    EXCEL_TAB_NAME,
    PRIMARY_KEY_FIELD,
    REQUEST_ID_FIELD,
    add_columns_to_right,
    get_row_identifiers_for_cosmos_matching,
)

from cdb import CDB

# --- Application-level imports ---
from config import Config

logger = logging.getLogger(__name__)
if not logging.getLogger().handlers:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(name)s - %(message)s")

COSMOS_DOC_PROCESS_DETAILS_FIELD = "process_details"

PROGRESS_UPDATE_INTERVAL = 100
COSMOS_BATCH_SIZE = 200  # Number of UIDs per IN clause (can be higher for UIDs than full doc IDs)


async def main():
    logger.info(
        f"Starting Cosmos DB results to Excel (batch by UID, filter by RequestID) for: '{EXCEL_FILE_PATH}', Sheet:"
        f" '{EXCEL_TAB_NAME}'"
    )

    main_app_config = Config()
    cdb_client = CDB(config=main_app_config)
    excel_wb = None

    logger.info("Cosmos DB client initialized.")

    try:
        row_identifiers_to_match = get_row_identifiers_for_cosmos_matching(
            file_path=EXCEL_FILE_PATH,
            sheet_name=EXCEL_TAB_NAME,
            pk_excel_header=PRIMARY_KEY_FIELD,
            request_id_excel_header=REQUEST_ID_FIELD,  # Still read RequestID for filtering
        )

        if not row_identifiers_to_match:
            logger.warning("No row identifiers found. Exiting.")
            return 0

        total_rows_to_process = len(row_identifiers_to_match)
        logger.info(f"Found {total_rows_to_process} Excel rows to process.")

        # --- Prepare data for batched UID queries ---
        # Map excel_row_num to its full row_info for later filtering and updates
        excel_row_info_map: Dict[int, Dict[str, Any]] = {
            row_info["excel_row_num"]: row_info for row_info in row_identifiers_to_match
        }

        # Get unique UIDs to query (pk_value from Excel, which is IVCE_DTL_UID)
        unique_uids_to_query_str = list({row_info["pk_value"] for row_info in row_identifiers_to_match if row_info["pk_value"]})

        # Convert UIDs to integers for Cosmos query, handling potential errors
        uids_for_cosmos_query_int: List[int] = []
        for uid_str in unique_uids_to_query_str:
            try:
                uids_for_cosmos_query_int.append(int(uid_str))
            except ValueError:
                logger.warning(
                    f"Invalid UID format '{uid_str}' found in Excel data; cannot convert to int. Skipping this UID for query."
                )

        if not uids_for_cosmos_query_int:
            logger.warning("No valid UIDs to query in Cosmos DB after filtering. Exiting data fetch.")
            all_fetched_cosmos_docs: List[Dict[str, Any]] = []
        else:
            logger.info(f"Will query Cosmos DB for {len(uids_for_cosmos_query_int)} unique UIDs.")

            # --- Fetch all potentially relevant Cosmos docs by UID in batches ---
            # Store all fetched docs in a list first
            all_fetched_cosmos_docs = []

            for i in range(0, len(uids_for_cosmos_query_int), COSMOS_BATCH_SIZE):
                batch_uids_int = uids_for_cosmos_query_int[i : i + COSMOS_BATCH_SIZE]

                # UIDs are numbers in Cosmos, so no quotes in the IN clause
                uids_str_for_in_clause = [str(uid) for uid in batch_uids_int]
                condition = f"c.invoice_details_from_rpa.IVCE_DTL_UID IN ({','.join(uids_str_for_in_clause)})"
                query_text = f"WHERE {condition}"

                batch_num = i // COSMOS_BATCH_SIZE + 1
                total_batches = (len(uids_for_cosmos_query_int) + COSMOS_BATCH_SIZE - 1) // COSMOS_BATCH_SIZE
                logger.info(
                    f"Fetching Cosmos batch {batch_num}/{total_batches} by UID. Condition snippet: ... IN"
                    f" ({len(uids_str_for_in_clause)} UIDs)"
                )

                try:
                    docs_in_batch = cdb_client.get_documents(container=cdb_client.ai_logs_container, where_condition=query_text)
                    all_fetched_cosmos_docs.extend(docs_in_batch)
                    logger.info(f"Batch {batch_num}: Fetched {len(docs_in_batch)} documents.")
                except Exception as e:
                    logger.error(f"Error in Cosmos batch UID query (Batch {batch_num}): {e}", exc_info=True)

        logger.info(f"Total documents fetched from Cosmos DB based on UIDs: {len(all_fetched_cosmos_docs)}")

        # --- Dynamically determine Excel columns from fetched data, ordered by stage ---
        # Key: sub_stage_name, Value: (stage_number, sub_stage_code) tuple for sorting
        stage_info_map: Dict[str, tuple] = {}
        for doc in all_fetched_cosmos_docs:
            process_details = doc.get(COSMOS_DOC_PROCESS_DETAILS_FIELD, {})
            if isinstance(process_details, dict):
                for stage_info in process_details.values():
                    if isinstance(stage_info, dict):
                        sub_stage_name = stage_info.get("sub_stage_name")
                        stage_number_val = stage_info.get("stage_number")
                        sub_stage_code_val = stage_info.get("sub_stage_code")

                        if sub_stage_name and stage_number_val is not None and sub_stage_code_val is not None:
                            try:
                                # Convert to numeric types for proper sorting
                                sort_key = (int(stage_number_val), float(sub_stage_code_val))
                                # If we haven't seen this sub_stage before, or if its new sort key is smaller, store it.
                                if sub_stage_name not in stage_info_map or sort_key < stage_info_map[sub_stage_name]:
                                    stage_info_map[sub_stage_name] = sort_key
                            except (ValueError, TypeError):
                                logger.warning(
                                    f"Could not parse stage_number/sub_stage_code for '{sub_stage_name}'. "
                                    "Skipping for column ordering."
                                )

        # Sort the collected stage names based on their sorting key (the tuple value)
        sorted_stage_items = sorted(stage_info_map.items(), key=lambda item: item[1])
        target_cosmos_data_excel_cols = [item[0] for item in sorted_stage_items]

        logger.info(
            f"Discovered and ordered {len(target_cosmos_data_excel_cols)} unique sub-stage names to use as columns: "
            f"{target_cosmos_data_excel_cols}"
        )

        # --- Prepare Excel for writing ---
        excel_wb = openpyxl.load_workbook(EXCEL_FILE_PATH)
        excel_ws = excel_wb[EXCEL_TAB_NAME]

        current_excel_headers = [str(cell.value).strip() for cell in excel_ws[1] if cell.value is not None]
        new_excel_headers_to_add = [
            col_name for col_name in target_cosmos_data_excel_cols if col_name not in current_excel_headers
        ]

        if new_excel_headers_to_add:
            logger.info(f"Adding new columns to Excel: {new_excel_headers_to_add}")
            add_columns_to_right(excel_ws, new_excel_headers_to_add)
            # Refresh headers list after adding new ones
            current_excel_headers = [str(cell.value).strip() for cell in excel_ws[1] if cell.value is not None]

        excel_col_name_to_idx_map: Dict[str, int] = {
            col_name: current_excel_headers.index(col_name) + 1 for col_name in target_cosmos_data_excel_cols
        }

        # --- Organize fetched docs for efficient lookup and filtering ---
        # Key: UID (string), Value: List of Cosmos docs with that UID
        cosmos_docs_by_uid_map: Dict[str, List[Dict[str, Any]]] = {}
        for doc in all_fetched_cosmos_docs:
            uid_from_doc_obj = doc.get("invoice_details_from_rpa", {}).get("IVCE_DTL_UID")
            if uid_from_doc_obj is not None:
                uid_str = str(uid_from_doc_obj)
                if uid_str not in cosmos_docs_by_uid_map:
                    cosmos_docs_by_uid_map[uid_str] = []
                cosmos_docs_by_uid_map[uid_str].append(doc)

        # --- Match and update Excel with fetched and filtered data ---
        logger.info("Matching fetched Cosmos data to Excel rows and updating sheet...")
        updated_excel_rows_count = 0

        for current_excel_row_num_processed, (excel_row_num, excel_row_data) in enumerate(excel_row_info_map.items()):
            excel_pk_uid = excel_row_data["pk_value"]  # This is the UID string from Excel
            excel_request_id = excel_row_data["request_id_value"]  # Optional RequestID from Excel

            matching_cosmos_doc: Optional[Dict[str, Any]] = None

            # Get candidate docs from our map using the Excel UID
            candidate_docs_for_uid = cosmos_docs_by_uid_map.get(excel_pk_uid, [])

            if candidate_docs_for_uid:
                if excel_request_id:
                    # Filter candidates by request_id if available
                    for candidate_doc in candidate_docs_for_uid:
                        cosmos_doc_request_id = candidate_doc.get("request_details", {}).get("id")
                        cosmos_doc_full_id = candidate_doc.get("id")
                        expected_doc_full_id = f"{excel_request_id}-{excel_pk_uid}"

                        if cosmos_doc_full_id == expected_doc_full_id:
                            matching_cosmos_doc = candidate_doc
                            break  # Exact match found by full ID
                        elif cosmos_doc_request_id == excel_request_id:
                            matching_cosmos_doc = candidate_doc
                    if not matching_cosmos_doc and candidate_docs_for_uid:
                        logger.debug(
                            f"Excel row {excel_row_num}: UID {excel_pk_uid} found {len(candidate_docs_for_uid)} doc(s) in Cosmos,"
                            f" but none matched Excel RequestID '{excel_request_id}'."
                        )
                else:
                    # No Excel RequestID, take the first candidate doc for this UID
                    matching_cosmos_doc = candidate_docs_for_uid[0]
                    if len(candidate_docs_for_uid) > 1:
                        logger.debug(
                            f"Excel row {excel_row_num}: UID {excel_pk_uid} has multiple ({len(candidate_docs_for_uid)}) Cosmos"
                            " docs. Using the first one as no Excel RequestID was provided for filtering."
                        )

            # First, clear all target columns for the current row to handle re-runs
            for col_name in target_cosmos_data_excel_cols:
                col_idx = excel_col_name_to_idx_map.get(col_name)
                if col_idx and excel_ws.cell(row=excel_row_num, column=col_idx).value is not None:
                    excel_ws.cell(row=excel_row_num, column=col_idx).value = None

            # Write data to Excel if a match was found
            if matching_cosmos_doc:
                process_details_from_doc = matching_cosmos_doc.get(COSMOS_DOC_PROCESS_DETAILS_FIELD, {})
                if isinstance(process_details_from_doc, dict):
                    for stage_data in process_details_from_doc.values():
                        if isinstance(stage_data, dict) and "sub_stage_name" in stage_data:
                            sub_stage_name = stage_data["sub_stage_name"]
                            if sub_stage_name in excel_col_name_to_idx_map:
                                excel_col_idx_to_write = excel_col_name_to_idx_map[sub_stage_name]
                                value_to_write_in_excel = json.dumps(stage_data)
                                excel_ws.cell(row=excel_row_num, column=excel_col_idx_to_write).value = value_to_write_in_excel

                updated_excel_rows_count += 1

            if (current_excel_row_num_processed + 1) % PROGRESS_UPDATE_INTERVAL == 0:
                logger.info(
                    f"Matched and prepared updates for {current_excel_row_num_processed + 1}/{total_rows_to_process} Excel rows"
                    " in memory..."
                )

        logger.info(f"Finished in-memory Excel updates. {updated_excel_rows_count} rows had Cosmos data populated.")
        logger.info("Attempting to save the Excel file...")
        excel_wb.save(EXCEL_FILE_PATH)
        logger.info(f"Successfully updated E2E Excel file: '{EXCEL_FILE_PATH}'")

        return 0

    except Exception as e:
        logger.error(f"An unexpected error occurred: {e}", exc_info=True)
        return 1
    finally:
        if excel_wb:
            excel_wb.close()


if __name__ == "__main__":
    exit_code = asyncio.run(main())
    sys.exit(exit_code)
